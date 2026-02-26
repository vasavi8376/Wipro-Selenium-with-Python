from openpyxl import load_workbook


class ExcelReader:

    def __init__(self, file_path, sheet_name):
        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook[sheet_name]

    def get_row_count(self):
        return self.sheet.max_row

    def get_column_count(self):
        return self.sheet.max_column

    def get_cell_data(self, row, column):
        return self.sheet.cell(row=row, column=column).value

